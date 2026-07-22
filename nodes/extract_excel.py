from openpyxl import load_workbook
from utils.ocr_utils import ocr_image


def extract_excel(state):
    path = state["file_path"]
    wb = load_workbook(path, data_only=True)

    text = ""
    for sheet in wb.worksheets:
        text += f"\n=== Sheet: {sheet.title} ===\n"

        rows = []
        for row in sheet.iter_rows(values_only=True):
            row_values = ["" if cell is None else str(cell) for cell in row]
            rows.append(" | ".join(row_values))

        if rows:
            text += "\n".join(rows) + "\n"

        images = getattr(sheet, '_images', [])
        if images:
            text += "\n[Excel images OCR] \n"
            for image_index, image_obj in enumerate(images, start=1):
                try:
                    image = image_obj.image
                    text += f"\n--- Image {image_index} ---\n"
                    text += ocr_image(image) + "\n"
                except Exception:
                    continue

    state["raw_text"] = text
    return state